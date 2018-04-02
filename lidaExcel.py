import openpyxl
import random

#biblioteca para lidar com arquivos no excel
def leMatriz(arquivo,aba):#arquivo xlsx, nome da aba 
    wb = openpyxl.load_workbook(arquivo)#wb recebe arquivo xlsx
    sheet = wb[aba]#sheet
    nLinhas = sheet.max_row
    nColunas = sheet.max_column
    matriz = []
    alfabeto = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    posAlfa = 0

    for rw in range(1,nLinhas + 1): #primeira linha do excel comeca em 1
        lista = []
        
        while posAlfa < nColunas:#copio uma linha inteira do excel para a lista
            lista.append(sheet[alfabeto[posAlfa] + str(rw)].value)
            posAlfa = posAlfa + 1
        #adiciono a linha na matriz e zero alfa pra ir para outra linha
        
        matriz.append(lista)
        posAlfa = 0
    return matriz #essa matriz tera os dados do excel



def imprimeMatriz(matriz):
    nLinhas = len(matriz)
    
    for i in range(0,nLinhas):
        print(matriz[i])

def geraMissingValues(matriz,porcentagem): #gera na matriz uma quantidade de buracos de acordo com a porcentagem
    
    qtdLinhas = len(matriz)
    qtdColunas = len(matriz[0])
    listaPosicoes = []
#    matrizMissingValues = copiaMatriz(matriz)

    N = qtdLinhas*qtdColunas #Numero total de registros
    N = N*porcentagem #numero de registros a terem buracos
    x=1

    while(x<N): #varia de 1 ate quantidade de registros a terem buracos
        posicao = []
        linha = random.randint(0, qtdLinhas-1)
        coluna = random.randint(0, qtdColunas-1)

        if verificaPosicaoMatriz(listaPosicoes, linha, coluna) == False:
            posicao.append(linha)
            posicao.append(coluna)
            listaPosicoes.append(posicao)
            matriz[linha][coluna] = "MISSING"
            x = x+1
    
    return matriz

def verificaPosicaoMatriz(listaPosicoes, linha, coluna): #verifica se a posicao ja foi gerada randomicamente

    for i in range (0,len(listaPosicoes)):
        if listaPosicoes[i][0] == linha and listaPosicoes[i][1] == coluna:
            return True 

    return False


def gravaNovaAba(matrizGrava,arquivo,titulo):#matriz, arquivo a ser gravado, titulo da aba
    wb = openpyxl.load_workbook(arquivo)#carrega o arquivo excel em memoria?

    wb.create_sheet(title=titulo)
    nLinhas = len(matrizGrava)
    nColunas = len(matrizGrava[0])
    
    sheet = wb[titulo]
    alfabeto = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    for i in range(0,nLinhas):
        posAlfa = 0
        for j in range(0,nColunas):

            sheet[alfabeto[posAlfa] + str(i+1)].value = matrizGrava[i][j]
            posAlfa = posAlfa + 1

    wb.save(arquivo)