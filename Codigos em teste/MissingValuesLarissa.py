import openpyxl, os
#biblioteca para lidar com arquivos no excel
from openpyxl import Workbook
import random
import numpy
import math 

def leMatriz(arquivo,aba):
	wb = openpyxl.load_workbook(arquivo)
	sheet = wb[aba]
	nLinhas = sheet.max_row
	nColunas = sheet.max_column
	matriz = []
	alfabeto = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
	posAlfa = 0

	for rw in range(1,nLinhas + 1): #primeira linha do excel comeca em 1
		lista = []
		while posAlfa < nColunas:
            #copio uma linha inteira do excel para a lista
			lista.append(sheet[alfabeto[posAlfa] + str(rw)].value)
			posAlfa = posAlfa + 1
        #adiciono a linha na matriz e zero alfa pra ir para outra linha
		matriz.append(lista)
		posAlfa = 0
	return matriz



def imprimeMatriz(matriz):
	nLinhas = len(matriz)
	
	for i in range(0,nLinhas):
		print(matriz[i])

def geraMissingValues(matriz,porcentagem): #gera na matriz uma quantidade de buracos de acordo com a porcentagem
	
	qtdLinhas = len(matriz)
	qtdColunas = len(matriz[0])
	listaPosicoes = []
#	matrizMissingValues = copiaMatriz(matriz)

	N = qtdLinhas*qtdColunas #Numero total de registros
	N = N*porcentagem #numero de registros a terem buracos
	x=1

	while(x<N): #varia de 1 ate quantidade de registros a terem buracos
		posicao = []
		linha = random.randint(0, qtdLinhas-1)
		coluna = random.randint(0, qtdColunas-1)

		if verificaPosicaoMatriz(listaPosicoes, linha, coluna) == False:
			posicao.append(linha)
			posicao.append(coluna)
			listaPosicoes.append(posicao)
			matriz[linha][coluna] = "MISSING"
			x = x+1
	
	return matriz

def verificaPosicaoMatriz(listaPosicoes, linha, coluna): #verifica se a posicao ja foi gerada randomicamente

	for i in range (0,len(listaPosicoes)):
		if listaPosicoes[i][0] == linha and listaPosicoes[i][1] == coluna:
			return True 

	return False


def gravaNovaAba(matrizGrava,arquivo,titulo):
	wb = openpyxl.load_workbook(arquivo)

	wb.create_sheet(title=titulo)
	nLinhas = len(matrizGrava)
	nColunas = len(matrizGrava[0])
	
	sheet = wb[titulo]
	alfabeto = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
	for i in range(0,nLinhas):
		posAlfa = 0
		for j in range(0,nColunas):

			sheet[alfabeto[posAlfa] + str(i+1)].value = matrizGrava[i][j]
			posAlfa = posAlfa + 1

	wb.save(arquivo)


def imputationMedia(matriz):

    #itero sobre cada coluna crio uma media com valor 0 inicial
	for coluna in range(1,len(matriz[1])+1):
		media = 0
		qtdRegistros = 0
        #itero sobre as linhas
		for i in range(0,len(matriz)):
			if matriz[i][coluna-1] != "MISSING":
				media = media + int(matriz[i][coluna-1])
				qtdRegistros = qtdRegistros + 1

		media = media/qtdRegistros

		for i in range(0,len(matriz)):
			if matriz[i][coluna-1] == "MISSING":
				matriz[i][coluna-1] = media
def imputationMode(matriz):
    for coluna in range(1,len(matriz[1])+1):
        repete = 0
        matriz1[:len(matriz)] = matriz[:len(matriz)].replace(0, numpy.NaN)
    return matriz1
imputationMode()
print(matriz1)


def imputationDesvioPadrao(matriz):

	for coluna in range(1,len(matriz[1])+1): #paracadacoluna
		media = 0
		qtdRegistros = 0
		for i in range(0,len(matriz)):
			if matriz[i][coluna-1] != "MISSING":
				media = media + matriz[i][coluna-1]
				qtdRegistros = qtdRegistros + 1

		media = media/qtdRegistros

		desvioPadrao = 0
		somatorio = 0
		for i in range(0,len(matriz)):
			if matriz[i][coluna-1] != "MISSING":
				somatorio = somatorio + pow(float(float(matriz[i][coluna-1]) - float(media)),2.0)

		desvioPadrao = somatorio/qtdRegistros
		desvioPadrao = math.sqrt(desvioPadrao)

		for i in range(0,len(matriz)):
			if matriz[i][coluna-1] == "MISSING":
				matriz[i][coluna-1] = desvioPadrao

def main():
	
	matriz = leMatriz('wine.xlsx','wine')

	geraMissingValues(matriz,0.1) 
	gravaNovaAba(matriz,'wine.xlsx',"MV")
	
	imputationMedia(matriz)

	gravaNovaAba(matriz,'wine.xlsx',"Imputation Media")
		
	matriz = []
	matriz = leMatriz('wine.xlsx','MV')

	imputationDesvioPadrao(matriz)

	gravaNovaAba(matriz,'wine.xlsx',"Deviation")

	
main()